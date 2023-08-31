#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Wed Sep  8 22:36:42 2021.

@author: manab
"""
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.cell import SHEETRANGE_RE

from ._make_spPr_from_dict import _make_spPr_from_dict


def _make_constraint_chart(
    ws,
    title,
    place_row,
    place_col,
    start_row,
    start_col,
    len_timeline,
    len_bargraph,
    len_linegraph=1,
    y_axis_title="[MWh]",
    x_axis_title="time",
    graphical_prop={},
):
    """
    各時間帯の制約式状況を積み上げグラフで出力する.

    Parameters
    ----------
    ws : CLASS
        グラフを出力するシートのインスタンス
    title : STR
        グラフタイトル
    place_row : INT
        グラフの行位置
    place_col : INT
        グラフの列位置
    start_row : INT
        データ領域開始の行位置
    start_col : INT
        データ領域開始の列位置
    len_timeline : INT
        時系列の長さ
    len_bargraph : INT
        棒グラフの数
    len_linegraph : INT
        線グラフの数
    y_axis_title : STR
        Y軸タイトル
    x_axis_title : STR
        X軸タイトル
    graphical_prop: DICT
        グラフの塗りつぶし色を指定する辞書型リスト
    """
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.x_axis.tickLblPos = "low"

    data = Reference(
        ws,
        min_row=start_row,
        max_row=start_row + len_timeline,
        min_col=start_col + 1,
        max_col=start_col + len_bargraph,
    )
    cats = Reference(ws, min_row=start_row + 1, min_col=start_col, max_row=len_timeline + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    for series in chart.series:
        m = SHEETRANGE_RE.match(series.tx.strRef.f)
        label = ws[m.group("cells")]._value
        if label == "_TRANSPARENT":
            series.tx.v = ""
            series.spPr.noFill = True
        if label in graphical_prop["bar"].keys():
            props = graphical_prop["bar"][label]
            _make_spPr_from_dict(series, props)

    if len_linegraph > 0:
        chart_l = LineChart()
        data_l = Reference(
            ws,
            min_row=start_row,
            max_row=start_row + len_timeline,
            min_col=start_col + len_bargraph + 1,
            max_col=start_col + len_bargraph + len_linegraph,
        )
        chart_l.add_data(data_l, titles_from_data=True)
        for series in chart_l.series:
            m = SHEETRANGE_RE.match(series.tx.strRef.f)
            label = ws[m.group("cells")]._value
            if label in graphical_prop["line"].keys():
                props = graphical_prop["line"][label]
                _make_spPr_from_dict(series, props)
        chart += chart_l

    chart.height = 13.5
    chart.width = 24
    place = ws.cell(row=place_row, column=place_col).coordinate
    ws.add_chart(chart, place)
