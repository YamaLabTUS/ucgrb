#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Wed Sep  8 22:36:42 2021.

@author: manab
"""
import math

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils.cell import SHEETRANGE_RE

from ._make_spPr_from_dict import _make_spPr_from_dict


def _make_object_function_chart(
    ws,
    title,
    place_row,
    place_col,
    start_row,
    start_col,
    len_timeline,
    len_elements,
    y_axis_title="[kJPY]",
    size=1,
    graphical_prop={},
):
    """
    目的関数（コスト、CO2排出量）を積み上げグラフと円グラフで出力する.

    Parameters
    ----------
    ws : CLASS
        グラフを出力するシートのインスタンス
    title : STR
        グラフタイトル
    place_row : INT
        グラフの行位置
    place_col : INT
        グラフの列位置
    start_row : INT
        データ領域開始の行位置
    start_col : INT
        データ領域開始の列位置
    len_timeline : INT
        時系列の長さ
    len_elements : INT
        要素の数
    y_axis_title: STR
        Y軸タイトル
    size: DOUBLE
        グラフのサイズ倍率
    graphical_prop: DICT
        グラフの塗りつぶし色を指定する辞書型リスト
    """
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = "time"
    chart.x_axis.tickLblPos = "low"

    data = Reference(
        ws,
        min_row=start_row,
        max_row=start_row + len_timeline,
        min_col=start_col + 1,
        max_col=start_col + len_elements,
    )
    cats = Reference(ws, min_row=start_row + 1, min_col=start_col, max_row=len_timeline + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    for series in chart.series:
        m = SHEETRANGE_RE.match(series.tx.strRef.f)
        label = ws[m.group("cells")]._value
        if label in graphical_prop["bar"].keys():
            props = graphical_prop["bar"][label]
            _make_spPr_from_dict(series, props)

    chart.height = 13.5 * size
    chart.width = 24 * size
    place = ws.cell(row=place_row, column=place_col).coordinate
    ws.add_chart(chart, place)

    chart_p = PieChart()
    chart_p.title = title

    data = Reference(
        ws,
        min_row=start_row + len_timeline + 1,
        min_col=start_col,
        max_col=start_col + len_elements,
    )
    cats = Reference(
        ws, min_row=start_row, min_col=start_col + 1, max_col=start_col + len_elements
    )

    chart_p.add_data(data, from_rows=True, titles_from_data=True)
    chart_p.set_categories(cats)
    chart_p.dLbls = DataLabelList(showVal=True, showCatName=True, showPercent=True)

    slices = [DataPoint(idx=i) for i in range(len_elements)]
    for i in range(len_elements):
        label = ws.cell(row=start_row, column=start_col + 1 + i).value
        if label in graphical_prop["bar"].keys():
            props = graphical_prop["bar"][label]
            _make_spPr_from_dict(slices[i], props)
    chart_p.series[0].data_points = slices

    chart_p.height = 13.5 * size
    chart_p.width = 24 * size
    place_p = ws.cell(row=place_row + math.ceil(29 * size), column=place_col).coordinate
    ws.add_chart(chart_p, place_p)
