#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Wed Sep  8 22:36:42 2021.

@author: manab
"""
import gurobipy as gp
from openpyxl.styles import Font

from ._append_col import _append_col
from ._make_constraint_chart import _make_constraint_chart
from ._make_object_function_chart import _make_object_function_chart


def about_all_area(ws, period_name, timeline, time_format, m, uc_data, uc_dicts):
    """
    全地域の需給状況を示すシートを作成する.

    Parameters
    ----------
    ws : CLASS
        結果を出力するシートのインスタンス
    period_name : STR
        表示対象の期間名称
    timeline : dataframe
        時系列
    time_format : STR
        時系列表示フォーマット
    m : CLASS
        Gurobiモデル
    uc_data : CLASS
        クラス「UCData」のインスタンス
    uc_dicts : CLASS
        クラス「UCDicts」のインスタンス

    """
    _header_col = ["Power Balance"] + list(timeline.keys().strftime(time_format))
    _append_col(ws, _header_col)
    _start_col = ws.max_column

    _value_col = ["Others"]
    for time in timeline:
        _value = uc_dicts.others_para["value"].sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    for g_type in uc_dicts.generation_type:
        _value_col = [g_type]
        for time in timeline:
            _value = uc_dicts.p.sum(time, "*", g_type, "*").getValue()
            _value_col.append(_value)
        _append_col(ws, _value_col)

    _value_col = ["Discharge of ESS"]
    for time in timeline:
        _value = uc_dicts.p_ess_d.sum(time, "*", "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Charge of ESS"]
    for time in timeline:
        _value = -uc_dicts.p_ess_c.sum(time, "*", "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["WF"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["WF_cap"][area] * uc_dicts.wf_para["output"][time, area]
                - uc_dicts.p_wf_suppr[time, area]
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["PV"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["PV_cap"][area] * uc_dicts.pv_para["output"][time, area]
                - uc_dicts.p_pv_suppr[time, area]
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Suppression of WF"]
    for time in timeline:
        _value = uc_dicts.p_wf_suppr.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Suppression of PV"]
    for time in timeline:
        _value = uc_dicts.p_pv_suppr.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Short"]
    for time in timeline:
        _value = uc_dicts.p_short.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Surplus"]
    for time in timeline:
        _value = -uc_dicts.p_surplus.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Demand"]
    for time in timeline:
        _value = uc_dicts.demand_para["value"].sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _make_constraint_chart(
        ws,
        "Power Balance in All Area on " + period_name,
        place_row=2,
        place_col=ws.max_column + 2,
        start_row=1,
        start_col=_start_col,
        len_timeline=len(timeline),
        len_bargraph=len(uc_dicts.generation_type) + 9,
        graphical_prop=uc_data.config["graphical_prop_for_xlsx_graph"],
    )

    ws.cell(column=ws.max_column + 1, row=1, value="")
    ws.insert_cols(ws.max_column, 12)

    # 目的関数(コスト)
    ws.insert_cols(ws.max_column, 2)  # 総コスト表示用に2列追加
    _header_col = ["Cost"] + list(timeline.keys().strftime(time_format)) + ["Sum"]
    _append_col(ws, _header_col)
    _start_col = ws.max_column
    total = 0

    for g_type in uc_dicts.n_and_t_generation_type:
        _value_col = ["Coef (" + g_type + ")"]
        _sum = 0
        for time in timeline:
            _value = gp.quicksum(
                uc_dicts.p[time, name, g_type, area].X
                * uc_dicts.generation_para["C_coef"][name, g_type, area]
                for name, g_type, area in uc_dicts.n_and_t_generation.select("*", g_type, "*")
            ).getValue()
            _value_col.append(_value)
            _sum += _value
        _value_col.append(_sum)
        _append_col(ws, _value_col)
        total += _sum

        _value_col = ["Intc (" + g_type + ")"]
        _sum = 0
        for time in timeline:
            _value = gp.quicksum(
                uc_dicts.u[time, name, g_type, area].X
                * uc_dicts.generation_para["C_intc"][name, g_type, area]
                for name, g_type, area in uc_dicts.n_and_t_generation.select("*", g_type, "*")
            ).getValue()
            _value_col.append(_value)
            _sum += _value
        _value_col.append(_sum)
        _append_col(ws, _value_col)
        total += _sum

        _value_col = ["Start Up (" + g_type + ")"]
        _sum = 0
        for time in timeline:
            _value = gp.quicksum(
                uc_dicts.su[time, name, g_type, area].X
                * uc_dicts.generation_para["C_startup"][name, g_type, area]
                for name, g_type, area in uc_dicts.n_and_t_generation.select("*", g_type, "*")
            ).getValue()
            _value_col.append(_value)
            _sum += _value
        _value_col.append(_sum)
        _append_col(ws, _value_col)
        total += _sum

    _value_col = ["Short"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.p_short[time, area].X * uc_dicts.area_para["C_short"][area]
            for area in uc_dicts.area
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Surplus"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.p_surplus[time, area].X * uc_dicts.area_para["C_surplus"][area]
            for area in uc_dicts.area
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Suppression of PV"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.p_pv_suppr[time, area].X * uc_dicts.area_para["C_PV_suppr"][area]
            for area in uc_dicts.area
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Suppression of WF"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.p_wf_suppr[time, area].X * uc_dicts.area_para["C_WF_suppr"][area]
            for area in uc_dicts.area
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Short of Tertiary (UP)"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.p_tert_up_short[time, area].X * uc_dicts.area_para["C_Tert_short"][area]
            for area in uc_dicts.area
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Short of Tertiary (DOWN)"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.p_tert_down_short[time, area].X * uc_dicts.area_para["C_Tert_short"][area]
            for area in uc_dicts.area
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["ESS Short Penalty"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.e_ess_short[time, ess, area].X * uc_dicts.ess_para["C_ess_short"][ess, area]
            for ess, area in uc_dicts.ess
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["ESS Surplus Penalty"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.e_ess_surplus[time, ess, area].X
            * uc_dicts.ess_para["C_ess_surplus"][ess, area]
            for ess, area in uc_dicts.ess
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    total_wo_tie = total
    _value_col = ["Tie Line Used Penalty"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.tie_para["C_tie_penalty"][name, f, t]
            * (uc_dicts.p_tie_f[time, name, f, t].X + uc_dicts.p_tie_c[time, name, f, t].X)
            for name, f, t in uc_dicts.tie
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Tie Line Used Penalty (GF&LFC, Up)"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.tie_para["C_tie_penalty_GF_LFC_UP"][name, f, t]
            * (
                uc_dicts.p_tie_gf_lfc_up_f[time, name, f, t].X
                + uc_dicts.p_tie_gf_lfc_up_c[time, name, f, t].X
            )
            for name, f, t in uc_dicts.tie
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Tie Line Used Penalty (GF&LFC, Down)"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.tie_para["C_tie_penalty_GF_LFC_DOWN"][name, f, t]
            * (
                uc_dicts.p_tie_gf_lfc_down_f[time, name, f, t].X
                + uc_dicts.p_tie_gf_lfc_down_c[time, name, f, t].X
            )
            for name, f, t in uc_dicts.tie
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Tie Line Used Penalty (Tert, Up)"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.tie_para["C_tie_penalty_Tert_UP"][name, f, t]
            * (
                uc_dicts.p_tie_tert_up_f[time, name, f, t].X
                + uc_dicts.p_tie_tert_up_c[time, name, f, t].X
            )
            for name, f, t in uc_dicts.tie
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    _value_col = ["Tie Line Used Penalty (Tert, Down)"]
    _sum = 0
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.tie_para["C_tie_penalty_Tert_DOWN"][name, f, t]
            * (
                uc_dicts.p_tie_tert_down_f[time, name, f, t].X
                + uc_dicts.p_tie_tert_down_c[time, name, f, t].X
            )
            for name, f, t in uc_dicts.tie
        ).getValue()
        _value_col.append(_value)
        _sum += _value
    _value_col.append(_sum)
    _append_col(ws, _value_col)
    total += _sum

    ws.cell(column=_start_col - 2, row=1, value="Total Cost [kJPY]")
    ws.cell(column=_start_col - 2, row=2, value=total_wo_tie)
    ws.cell(
        column=_start_col - 2,
        row=4,
        value="Total Cost with Used Penalty of Tie Line [kJPY]",
    )
    ws.cell(column=_start_col - 2, row=5, value=total)

    cordinate = ws.cell(column=_start_col - 2, row=2).coordinate
    ws[cordinate].font = Font(bold=True)

    _make_object_function_chart(
        ws,
        "Cost in All Area on " + period_name,
        place_row=2,
        place_col=ws.max_column + 2,
        start_row=1,
        start_col=_start_col,
        len_timeline=len(timeline),
        len_elements=len(uc_dicts.n_and_t_generation_type) * 3 + 13,
        size=1.5,
        graphical_prop=uc_data.config["graphical_prop_for_xlsx_graph"],
    )

    ws.cell(column=ws.max_column + 1, row=1, value="")
    ws.insert_cols(ws.max_column, 18)

    # CO2排出量
    ws.insert_cols(ws.max_column, 2)  # 総排出量表示用に2列追加
    _header_col = ["CO2 Emission"] + list(timeline.keys().strftime(time_format)) + ["Sum"]
    _append_col(ws, _header_col)
    _start_col = ws.max_column
    total = 0

    num_g_type = 0
    for g_type in uc_dicts.generation_type:
        if g_type in ["HYDRO", "NUCL"]:
            continue
        num_g_type = num_g_type + 1
        _value_col = ["Coef (" + g_type + ")"]
        _sum = 0
        for time in timeline:
            _value = gp.quicksum(
                uc_dicts.p[time, name, g_type, area].X
                * uc_dicts.generation_para["C_coef_CO2"][name, g_type, area]
                for name, g_type, area in uc_dicts.generation.select("*", g_type, "*")
            ).getValue()
            _value_col.append(_value)
            _sum += _value
        _value_col.append(_sum)
        _append_col(ws, _value_col)
        total += _sum

        _value_col = ["Intc (" + g_type + ")"]
        _sum = 0
        for time in timeline:
            _value = gp.quicksum(
                uc_dicts.u[time, name, g_type, area].X
                * uc_dicts.generation_para["C_intc_CO2"][name, g_type, area]
                for name, g_type, area in uc_dicts.n_and_t_generation.select("*", g_type, "*")
            ).getValue()
            _value_col.append(_value)
            _sum += _value
        _value_col.append(_sum)
        _append_col(ws, _value_col)
        total += _sum

        _value_col = ["Start Up (" + g_type + ")"]
        _sum = 0
        for time in timeline:
            _value = gp.quicksum(
                uc_dicts.su[time, name, g_type, area].X
                * uc_dicts.generation_para["C_startup_CO2"][name, g_type, area]
                for name, g_type, area in uc_dicts.n_and_t_generation.select("*", g_type, "*")
            ).getValue()
            _value_col.append(_value)
            _sum += _value
        _value_col.append(_sum)
        _append_col(ws, _value_col)
        total += _sum

    ws.cell(column=_start_col - 2, row=1, value="Total CO2 Emission [tCO2]")
    ws.cell(column=_start_col - 2, row=2, value=total)

    cordinate = ws.cell(column=_start_col - 2, row=2).coordinate
    ws[cordinate].font = Font(bold=True)

    _make_object_function_chart(
        ws,
        "CO2 Emission in All Area on " + period_name,
        place_row=2,
        place_col=ws.max_column + 2,
        start_row=1,
        start_col=_start_col,
        len_timeline=len(timeline),
        len_elements=num_g_type * 3,
        y_axis_title="[tCO2]",
        graphical_prop=uc_data.config["graphical_prop_for_xlsx_graph"],
    )

    ws.cell(column=ws.max_column + 1, row=1, value="")
    ws.insert_cols(ws.max_column, 12)

    # PV出力時系列
    _header_col = ["PV"] + list(timeline.keys().strftime(time_format))
    _append_col(ws, _header_col)
    _start_col = ws.max_column

    _value_col = ["_TRANSPARENT"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["PV_cap"][area] * uc_dicts.pv_para["output"][time, area]
                - uc_dicts.p_pv_suppr[time, area].X
                - uc_dicts.p_pv_tert_down[time, area].X
                - uc_dicts.p_pv_gf_lfc_down[time, area].X
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Tertiary (Down)"]
    for time in timeline:
        _value = uc_dicts.p_pv_tert_down.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["GF&LFC (Down)"]
    for time in timeline:
        _value = uc_dicts.p_pv_gf_lfc_down.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["GF&LFC (Up)"]
    for time in timeline:
        _value = uc_dicts.p_pv_gf_lfc_up.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Tertiary (Up)"]
    for time in timeline:
        _value = uc_dicts.p_pv_tert_up.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["PV Net"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["PV_cap"][area] * uc_dicts.pv_para["output"][time, area]
                - uc_dicts.p_pv_suppr[time, area].X
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["PV Output"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["PV_cap"][area] * uc_dicts.pv_para["output"][time, area]
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Reserve limit (Up)"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["PV_cap"][area] * uc_dicts.pv_para["output"][time, area]
                - uc_dicts.p_pv_suppr[time, area].X
                * (1 - uc_dicts.area_para["R_PV_res_UP"][area] / 100)
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Reserve limit (Down)"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                (
                    uc_dicts.area_para["PV_cap"][area] * uc_dicts.pv_para["output"][time, area]
                    - uc_dicts.p_pv_suppr[time, area].X
                )
                * (1 - uc_dicts.area_para["R_PV_res_DOWN"][area] / 100)
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _make_constraint_chart(
        ws,
        "PV in All Area on " + period_name,
        place_row=2,
        place_col=ws.max_column + 2,
        start_row=1,
        start_col=_start_col,
        len_timeline=len(timeline),
        len_bargraph=5,
        len_linegraph=4,
        graphical_prop=uc_data.config["graphical_prop_for_xlsx_graph"],
    )

    ws.cell(column=ws.max_column + 1, row=1, value=" ")
    ws.insert_cols(ws.max_column, 12)

    # WF出力時系列
    _header_col = ["WF"] + list(timeline.keys().strftime(time_format))
    _append_col(ws, _header_col)
    _start_col = ws.max_column

    _value_col = ["_TRANSPARENT"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["WF_cap"][area] * uc_dicts.wf_para["output"][time, area]
                - uc_dicts.p_wf_suppr[time, area].X
                - uc_dicts.p_wf_tert_down[time, area].X
                - uc_dicts.p_wf_gf_lfc_down[time, area].X
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Tertiary (Down)"]
    for time in timeline:
        _value = uc_dicts.p_wf_tert_down.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["GF&LFC (Down)"]
    for time in timeline:
        _value = uc_dicts.p_wf_gf_lfc_down.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["GF&LFC (Up)"]
    for time in timeline:
        _value = uc_dicts.p_wf_gf_lfc_up.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Tertiary (Up)"]
    for time in timeline:
        _value = uc_dicts.p_wf_tert_up.sum(time, "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["WF Net"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["WF_cap"][area] * uc_dicts.wf_para["output"][time, area]
                - uc_dicts.p_wf_suppr[time, area].X
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["WF Output"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["WF_cap"][area] * uc_dicts.wf_para["output"][time, area]
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Reserve limit (Up)"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                uc_dicts.area_para["WF_cap"][area] * uc_dicts.wf_para["output"][time, area]
                - uc_dicts.p_wf_suppr[time, area].X
                * (1 - uc_dicts.area_para["R_WF_res_UP"][area] / 100)
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Reserve limit (Down)"]
    for time in timeline:
        if time in uc_dicts.timeline:
            _value = gp.quicksum(
                (
                    uc_dicts.area_para["WF_cap"][area] * uc_dicts.wf_para["output"][time, area]
                    - uc_dicts.p_wf_suppr[time, area].X
                )
                * (1 - uc_dicts.area_para["R_WF_res_DOWN"][area] / 100)
                for area in uc_dicts.area
            ).getValue()
        else:
            _value = 0
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _make_constraint_chart(
        ws,
        "WF in All Area on " + period_name,
        place_row=2,
        place_col=ws.max_column + 2,
        start_row=1,
        start_col=_start_col,
        len_timeline=len(timeline),
        len_bargraph=5,
        len_linegraph=4,
        graphical_prop=uc_data.config["graphical_prop_for_xlsx_graph"],
    )

    ws.cell(column=ws.max_column + 1, row=1, value=" ")
    ws.insert_cols(ws.max_column, 12)

    # エネルギー貯蔵システム
    _header_col = ["ESS"] + list(timeline.keys().strftime(time_format))
    _append_col(ws, _header_col)
    _start_col = ws.max_column

    _value_col = ["Discharge of ESS"]
    for time in timeline:
        _value = uc_dicts.p_ess_d.sum(time, "*", "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Charge of ESS"]
    for time in timeline:
        _value = -uc_dicts.p_ess_c.sum(time, "*", "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Energy"]
    for time in timeline:
        _value = uc_dicts.e_ess.sum(time, "*", "*").getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Capacity"]
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.ess_para["E_CAP"][name, area] for name, area in uc_dicts.ess
        ).getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Max"]
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.ess_para["E_CAP"][name, area]
            * uc_dicts.ess_para["E_R_MAX"][name, area]
            / 100
            for name, area in uc_dicts.ess
        ).getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Min"]
    for time in timeline:
        _value = gp.quicksum(
            uc_dicts.ess_para["E_CAP"][name, area]
            * uc_dicts.ess_para["E_R_MIN"][name, area]
            / 100
            for name, area in uc_dicts.ess
        ).getValue()
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _value_col = ["Energy Plan"]
    for time in timeline:
        if uc_data.config["set_e_ess_bc_constrs"] and time == uc_dicts.timeline[-1]:
            _value = gp.quicksum(
                uc_dicts.ess_para["E_CAP"][name, area]
                * uc_dicts.ess_para["E_R_base"][name, area]
                / 100
                for name, area in uc_dicts.ess
            ).getValue()
        elif (
            uc_data.config["set_e_ess_plan_constrs"]
            and time in uc_dicts.whole_timeline_ess_plan.values
        ):
            _value = gp.quicksum(
                uc_dicts.ess_para["E_CAP"][name, area]
                * uc_dicts.e_ess_plan_para["value"][time, name, area]
                / 100
                for name, area in uc_dicts.ess
            ).getValue()
        else:
            _value = ""
        _value_col.append(_value)
    _append_col(ws, _value_col)

    _make_constraint_chart(
        ws,
        "Energy Storage in All Area on " + period_name,
        place_row=2,
        place_col=ws.max_column + 2,
        start_row=1,
        start_col=_start_col,
        len_timeline=len(timeline),
        len_bargraph=2,
        len_linegraph=5,
        y_axis_title="[MWh]",
        graphical_prop=uc_data.config["graphical_prop_for_xlsx_graph"],
    )

    ws.cell(column=ws.max_column + 1, row=1, value=" ")
    ws.insert_cols(ws.max_column, 12)
