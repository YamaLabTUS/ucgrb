#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 27 14:02:03 2021.

@author: y_hcr_manabe
"""
import os
import zipfile

import gurobipy as gp
from openpyxl import Workbook

from .._make_dir import _make_dir
from ._make_xlsx_sheet import (
    _make_xlsx_sheet_about_all_area,
    _make_xlsx_sheet_about_area,
    _make_xlsx_sheet_about_ess,
    _make_xlsx_sheet_about_generation,
    _make_xlsx_sheet_about_shadow_price,
    _make_xlsx_sheet_about_tie,
)


def output_result(m, uc_data, uc_dicts, i):
    """
    最適化結果をxlsxファイル等で出力、保存する.

    Parameters
    ----------
    m : CLASS
        Gurobiモデル
    uc_data : CLASS
        クラス「UCData」のインスタンス
    uc_dicts : CLASS
        クラス「UCDicts」のインスタンス
    i : int
        最適化リスト中、現在何回目を実施しているかを示すインデックス
    """
    if "_export_dir" not in uc_data.config:
        uc_data.config["_export_dir"] = _make_dir(
            uc_data.config["result_dir"], uc_data.config["_identify_str"]
        )
    _export_json_file(m, uc_data, uc_dicts, i)
    _export_xlsx_file(m, uc_data, uc_dicts, i)


def _export_json_file(m, uc_data, uc_dicts, i):
    """
    最適化結果をjsonファイルに保存する.

    Parameters
    ----------
    m : CLASS
        Gurobiモデル
    uc_data : CLASS
        クラス「UCData」のインスタンス
    uc_dicts : CLASS
        クラス「UCDicts」のインスタンス
    i : int
        最適化リスト中、現在何回目を実施しているかを示すインデックス
    """

    if uc_data.config["export_json_file"] is False:
        return
    _opt = uc_data.config["rolling_opt_list"][i]
    _dir = _make_dir(str(uc_data.config["_export_dir"]), "json")
    filename = _opt["name"] + ".json"
    filepath = str(_dir / filename)

    m.write(filepath)

    zip_file = zipfile.ZipFile(filepath + ".zip", mode="w", compression=zipfile.ZIP_DEFLATED)
    zip_file.write(filepath, arcname=filename)
    os.remove(filepath)


def _export_xlsx_file(m, uc_data, uc_dicts, i):
    """
    最適化結果をxlsxファイルに保存する.

    Parameters
    ----------
    m : CLASS
        Gurobiモデル
    uc_data : CLASS
        クラス「UCData」のインスタンス
    uc_dicts : CLASS
        クラス「UCDicts」のインスタンス
    i : int
        最適化リスト中、現在何回目を実施しているかを示すインデックス
    """
    if uc_data.config["export_xlsx_file"] is False:
        return
    _opt = uc_data.config["rolling_opt_list"][i]

    # シャドープライス算出のために、MIPの場合に整数を固定して再度最適を実施する。
    if m.IsMIP == 0:
        m_Pi = m.copy()
    else:
        m_Pi = m.fixed()
        if "grb_FeasibilityTol_for_Pi_calc" in uc_data.config:
            m_Pi.Params.FeasibilityTol = uc_data.config["grb_FeasibilityTol_for_Pi_calc"]
        lf = "./" + str(uc_data.config["_export_dir"]) + "/log/" + _opt["name"] + "_PI.log"
        m_Pi.Params.LogFile = lf
        m_Pi.update()

        if uc_data.config["export_mps_file"]:
            filename = _opt["name"] + "_PI.mps"
            filepath = "./" + str(uc_data.config["_export_dir"]) + "/mps/" + filename

            m_Pi.write(filepath)

            zip_file = zipfile.ZipFile(
                filepath + ".zip", mode="w", compression=zipfile.ZIP_DEFLATED
            )
            zip_file.write(filepath, arcname=filename)
            os.remove(filepath)

        m_Pi.optimize()

        if uc_data.config["export_json_file"]:
            filename = _opt["name"] + "_PI.json"
            filepath = "./" + str(uc_data.config["_export_dir"]) + "/json/" + filename

            m_Pi.Params.JSONSolDetail = 1
            for c in m_Pi.getConstrs():
                c.CTag = c.ConstrName
            m_Pi.update()
            m_Pi.write(filepath)

            zip_file = zipfile.ZipFile(
                filepath + ".zip", mode="w", compression=zipfile.ZIP_DEFLATED
            )
            zip_file.write(filepath, arcname=filename)
            os.remove(filepath)

    filename = _opt["name"] + "_chart.xlsx"
    period_name = _opt["name"]
    _make_xlsx_book(
        filename,
        period_name,
        uc_dicts.timeline_pickup_in_result_file,
        "%Y/%m/%d %H:%M",
        m,
        m_Pi,
        uc_data,
        uc_dicts,
    )

    if uc_dicts.timeline_pickup_in_result_file.equals(uc_dicts.timeline) is False:
        filename = _opt["name"] + "_all_time_chart.xlsx"
        _make_xlsx_book(
            filename,
            period_name,
            uc_dicts.timeline,
            "%Y/%m/%d %H:%M",
            m,
            m_Pi,
            uc_data,
            uc_dicts,
        )


def _make_xlsx_book(filename, period_name, timeline, time_format, m, m_Pi, uc_data, uc_dicts):
    wb = Workbook()

    # 全地域の需給状況
    ws = wb.active
    ws.title = "total"

    _make_xlsx_sheet_about_all_area(ws, period_name, timeline, time_format, m, uc_data, uc_dicts)

    # 各地域の需給状況
    for area in uc_dicts.area:
        ws = wb.create_sheet(area)

        _make_xlsx_sheet_about_area(
            ws, period_name, timeline, time_format, area, m, uc_data, uc_dicts
        )

    # 各地域のシャドウプライス
    if uc_data.config["export_xlsx_file"]["shadow_price"] is True:
        ws = wb.create_sheet("shadow price")

        _make_xlsx_sheet_about_shadow_price(
            ws, period_name, timeline, time_format, m_Pi, uc_data, uc_dicts
        )

    # 各地域の大規模発電機の運用時系列
    if uc_data.config["export_xlsx_file"]["generation"] is True:
        for area in uc_dicts.area:
            ws = wb.create_sheet(area + "_generation")

            _make_xlsx_sheet_about_generation(
                ws, period_name, timeline, time_format, area, m, uc_data, uc_dicts
            )

    # 各地域のエネルギー貯蔵システムの運用時系列
    if uc_data.config["export_xlsx_file"]["ESS"] is True:
        for area in uc_dicts.area:
            ws = wb.create_sheet(area + "_ESS")

            _make_xlsx_sheet_about_ess(
                ws, period_name, timeline, time_format, area, m, uc_data, uc_dicts
            )

    # 各連系線の運用時系列
    if uc_data.config["export_xlsx_file"]["tie"] is True:
        for tie, f, t in uc_dicts.tie:
            ws = wb.create_sheet(tie)

            _make_xlsx_sheet_about_tie(
                ws, period_name, timeline, time_format, tie, f, t, m, uc_data, uc_dicts
            )

    wb.save(uc_data.config["_export_dir"] / filename)
    wb.close()
